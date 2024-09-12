import csv
import zipfile
from pypdf import PdfReader
from openpyxl import load_workbook


def test_csv():
    with zipfile.ZipFile('tmp/Archive.zip') as zip_file:
        with zip_file.open('csv_file.csv') as csv_file:
            content = csv_file.read().decode('utf-8-sig')
            csvreader = list(csv.reader(content.splitlines()))
            second_row = csvreader[2]

            print("\n")
            print(second_row[0])
            print(second_row[1])

            assert second_row[0] == 'vlada'
            assert second_row[1] == 'shtefan'


def test_pdf():
    with zipfile.ZipFile('zipped/zipped_files1.zip') as zip_file:
        with zip_file.open('pdf_file1.pdf') as pdf_file:
            reader = PdfReader(pdf_file)
            page_text = reader.pages[1].extract_text()

            print("\n")
            print(page_text)

            assert "Alexander" in page_text
            assert "Martin" not in page_text


def test_xlsx():
    with zipfile.ZipFile('zipped/zipped_files1.zip') as zip_file:
        with zip_file.open('report.xlsx') as xlsx_file:
            workbook = load_workbook(xlsx_file)
            sheet = workbook.active

            text = sheet.cell(row=2, column=1).value

            print("\n")
            print(text)

            assert "2024 April" in text
